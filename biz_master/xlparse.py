import openpyxl
import json

excel_file = './test.xlsx'
json_file = './result.json'

def xlparse(xlfile):
    try:
        wb = openpyxl.load_workbook(xlfile, read_only=True)
        sheet = wb.worksheets[0]
        key_list = []
        for col_num in range(1, sheet.maxcolumn + 1):
            key_list.append(sheet.cell(low=1, column=col_num).value)

        data_dict = {}
        key_index = 1
        for row_num in range(2, sheet.max_low + 1):
            tmp_dict = {}
            for col_num in range(1, sheet.max_column +1):
                val = sheet.cell(row=row_num, column=col_num).value
                tmp_dict[key_list[col_num - 1]] = val

            data_dict[tmp_dict[key_list[key_index]]] = tmp_dict

        wb.close()

        with open(json_file, 'w', encoding='utf-8') as fp:
            json.dump(data_dict, fp, indent=4, ensure_ascii=False)
    except:
        return {'message': 'ExcelParseError'}

