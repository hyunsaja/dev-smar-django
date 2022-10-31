# from django.contrib.auth.models import User
# from .xlparse import xlparse
# from datetime import datetime, date, timedelta
# from django.utils.dateformat import DateFormat
# from rest_framework.decorators import api_view

import json
import math
import time

import numpy
from rest_framework.response import Response
from django.http import HttpResponse
from django.shortcuts import render
from rest_framework import status
from rest_framework.views import APIView

# ==============================================================================
# 마킹 머신 공통 -----------------------------------------------------------------
# ==============================================================================
# 미주 마킹기 : 1
# machine_id = 1  # 추후 자동으로 처리

# Error = {}
# Error['message'] = 'Error'
# Ok = {}
# Ok['message'] = 'Ok'

Ok = 'Ok'
Error = 'Error'

# ==============================================================================
# public lib
# ==============================================================================
from .models import AutoMarkMachine, AutoPressMachine, RpcagMachine
from .serializers import AutoMarkSerializer, AutoPressSerializer, RpcagSerializer
from openpyxl import load_workbook


# ==============================================================================
# mark_machine excel upload ----------------------------------------------------
# ==============================================================================


class UploadExcelData(APIView):
    def get(self, request, *args, **kwargs):
        print("======================")
        print(request)
       # print(request.GET['extra_machineID'])
        print(request.GET.get('aData'))

        print("======================")
        return HttpResponse(None)

    def post(self, request, *args, **kwargs):
        # file = request.FILES['file_excel']
        # a = request.POST['extra_machineID']
        # b = request.POST['aData']
        # print(file)
        # print(a)
        # print(b)
        # return Response('Ok')

        try:
            file = request.FILES['file_excel']
            machine_id = request.data['extra_machineID']
            print(machine_id)
            # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
            wb = load_workbook(file, read_only=True)
            sheet = wb.worksheets[0]

            author = request.user.id
            key_list = ['temp', 'ship_no', 'por_no', 'seq_no', 'block_no', 'pcs_no', 'paint_code', 'lot_no']
            data_list = []
            low_count = 0
            for row in sheet.rows:
                if low_count == 0:
                    low_count = low_count + 1  # 제목행 pass
                    continue

                td = {}
                count = 0  # No 컬럼 건너뛰기
                #-------------------------------------
                for cell in row:
                    td[key_list[count]] = str(cell.value)
                    if td[key_list[count]] == 'None':
                        td[key_list[count]] = '-'
                    count = count + 1
                # -------------------------------------

                td['author'] = author
                td['machine_id'] = machine_id
                del td['temp']  # No 항목 삭제

                # 마크데이터 자동생성
                if td['lot_no'] != None:
                    td['mark_data'] = td['paint_code'] + ' ' + td['ship_no'] + ' ' + td['por_no'] + '-'\
                                    + td['seq_no'] + ' ' + td['block_no'] + ' ' + td['pcs_no'] + ' ' + td['lot_no']
                else:
                    td['mark_data'] = td['paint_code'] + ' ' + td['ship_no'] + ' ' + td['por_no'] + '-'\
                                    + td['seq_no'] + ' ' + td['block_no'] + ' ' + td['pcs_no']

                data_list.append(td)
            wb.close()

            serializer = AutoMarkSerializer(data=data_list, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(Ok)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except:
            return Response(Error)


# ==============================================================================
#  work 그룹 조회 ----------------------------------------------------------------
# ==============================================================================

class MworkDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID']) #codesys에 박아넣음
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)
            try:
                mark_list = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                           status=False)
            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

            tmp_dict = {}
            for obj in mark_list:  # 시퀀스의 피스 합계 구하는 파트
                if obj.work_select == True:
                    sel = 'o'
                else:
                    sel = 'x'
                group = obj.ship_no + ',' + obj.por_no + ',' + obj.seq_no + ',' + sel
                tmp_dict[group] = tmp_dict.get(group, 0) + obj.work_quantity

            mark_dict = {}
            list = []
            tmp = tmp_dict.keys()

            count = len(tmp)
            for k in tmp:
                mark_dict['count'] = str(count)
                k_list = k.split(',')
                mark_dict['ship_no'] = k_list[0]
                mark_dict['por_no'] = k_list[1]
                mark_dict['seq_no'] = k_list[2]
                mark_dict['work_quantity'] = str(tmp_dict[k])
                mark_dict['work_select'] = k_list[3]  # str 임
                list.append(mark_dict)
                mark_dict = {}  # 비워주지 않으면 마지막 데이터로 리스트가 다 채워짐

            return Response(list)
        except:
            return Response(Error)


# ==============================================================================
#  마킹 데이터 내려받기------------------------------------------------------------
# ==============================================================================

class MarkDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            try:
                markdatas = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True)[:100]
                count = markdatas.count()
                mark_list = []
                mark_dict = {}
                for m in markdatas:
                    mark_dict['count'] = str(count)
                    mark_dict['mark_data'] = m.mark_data
                    mark_dict['work_quantity'] = str(m.work_quantity)
                    mark_dict['worked_quantity'] = str(m.worked_quantity)
                    print(mark_dict)
                    mark_list.append(mark_dict)
                    mark_dict = {}

                return Response(mark_list)

            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)
        except:
            return Response(Error)


# ==============================================================================
#  마킹 그룹 개별 작업 선택 ----------------------------------------------------------------
# ==============================================================================

class MarkGselect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(json.dumps({'message':'Error'}))
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            ship_no = request.data['project']
            por_no = request.data['por']
            seq_no = request.data['seq']
            print(ship_no)

            try:
                select_fields = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                ship_no=ship_no,
                                                                por_no=por_no,
                                                                seq_no=seq_no)
                select_fields.update(work_select=True)
                return Response(Ok)

            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

        except:
            return Response(Error)

# ==============================================================================
#  마킹 그룹 전체 작업 선택 ----------------------------------------------------------------
# ==============================================================================

class MarkAselect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            try:
                select_fields = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                                status=False)
                select_fields.update(work_select=True)
                return Response(Ok)

            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

        except:
            return Response(Error)



# ==============================================================================
#  마킹 그룹 선택 해제 ----------------------------------------------------------------
# ==============================================================================

class MarkGcancle(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            try:
                select_fields = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                work_select=True)
                select_fields.update(work_select=False)
                return Response(Ok)

            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

        except:
            return Response(Error)


# ==============================================================================
#  마킹 실적 등록------------------------------------------------------------
# ==============================================================================

class MarkedData(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            mark_data = request.data['mark_data']

            try:
                markdatas = AutoMarkMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True,
                                                            mark_data=mark_data)
                for m in markdatas:
                    if m.work_quantity > m.worked_quantity:
                        m.worked_quantity = m.worked_quantity + 1
                        if m.work_quantity == m.worked_quantity:
                            m.status = True
                    else:
                        m.status = True
                    m.save()

                return Response(Ok)

            except AutoMarkMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)
        except:
            return Response(Error)



# ==============================================================================
# ==============================================================================
# robot_angle_cutting_machine ---- cutting view(get)
# ==============================================================================
# ==============================================================================

class RpcagCuttingView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            machine_id = int(request.GET['machineID'])

            view_data = request.GET['material']
            standard = request.GET['standard']

            cutlist = []
            try:
                cutdata = RpcagMachine.objects.get(
                    machine_id=machine_id,
                    status=False,
                    work_select=True,
                    standard=standard,
                    view_data=view_data)
                cutlist = cutdata.cutlist
                print(cutlist)
                standard = cutdata.standard
                print(standard)
                length = cutdata.length_dwg
            except AutoPressMachine.DoesNotExist:
                return Response(Error)

            return render(request,
                          'code_master/AutoPressMachine/press_cutting_view.html',
                          {'cutlist': cutlist, 'standard': standard,
                           'length': length})

        except:
            return Response('codeError')


# ==============================================================================
# robot_angle_cutting_machine ---- json upload
# ==============================================================================

class UploadRpcagJsonData(APIView):
    def post(self, request, *args, **kwargs):

        try:
            file = request.FILES['json_file']
            machine_id = int(request.data['extra_machineID'])
            author = request.user.id
            print(author)

            fielddata = {}
            fieldlist = []
            jsonfile = json.load(file)
            for fielddatas in jsonfile:  # 각 피스 단위로 추출됨: [{피스정보},{피스정보}]
                for key in fielddatas['part'].keys():  # 부재 단위로 추출됨: part : [{ea01:{},ea02:{},...}]
                    fielddata = fielddatas['part'][key]  # 부재 정보 담김
                    fielddata['author'] = author
                    fielddata['machine_id'] = machine_id
                    fielddata['view_data'] = fielddatas['Material'] + '~' + key  # key는 부재번호임

                    g_list = fielddatas['Material'] .split('~')
                    del g_list[-1]  # 피스번호 지움
                    count = len(g_list)
                    if count == 2:
                        group = g_list[0] + '~' + g_list[1] + '~' + fielddata['texture']
                    elif count == 3:
                        group = g_list[0] + '~' + g_list[1] + '~' + g_list[2] + '~' + fielddata['texture']
                    else:
                        return Response(Error)
                    fielddata['group_data'] = group
                    quantity = int(fielddatas['quantity']) * int(fielddata['quantity'])
                    fielddata['work_quantity'] = quantity
                    fielddata['length_dwg'] = int(fielddata['length'])
                    fielddata['length_cut'] = int(fielddata['length_real'])
                    fielddata['part_point'] = int(fielddata['part_point'])
                    del fielddata['creation_date']
                    del fielddata['quantity']
                    del fielddata['length']
                    del fielddata['length_real']

                    if RpcagMachine.objects.filter(view_data=fielddata["view_data"]).exists():
                        continue

                    m_kind = fielddata['standard'].split(' ')[0]
                    if m_kind != 'EA':
                        continue

                    # Ubolt 변환...
                    cutlist = fielddata['cutlist']
                    cutlist2 = []
                    for cut in cutlist:  # 가공 매크로 순회
                        # cuts는 {"CUT": [거리, 매크로(홀), 상수A(홀크기), 상수B(높이), 상수C, 상수D]}의 리스트
                        # U볼트 정보 ~ 홀(파이), 높이, 규격(A), Pitch, 홀체크(0,1),슈여부(0,1), 슈홀(파이), 슈높이
                        # {"CUT": ["100", "UBOLT0", "12", "22", "50A", "74", "1", "0", "0", "25"]}
                        if 'UBOLT' in cut['CUT'][1]:
                            start = {"CUT": [str(int(cut['CUT'][0]) - int(cut['CUT'][5]) / 2),
                                m_kind + cut['CUT'][1][-1] + '11', cut['CUT'][2], cut['CUT'][3], "0", "0"]}
                            cutlist2.append(start)
                            end = {"CUT": [str(int(cut['CUT'][0]) + int(cut['CUT'][5]) / 2),
                                           m_kind + cut['CUT'][1][-1] + '11',
                                           cut['CUT'][2], cut['CUT'][3],
                                           "0", "0"]}
                            cutlist2.append(end)
                            if cut['CUT'][7] == '1':
                                mid = {"CUT": [cut['CUT'][0], m_kind + cut['CUT'][1][-1] + '11',
                                               cut['CUT'][8], cut['CUT'][9], "0", "0"]}
                                cutlist2.append(mid)
                        else:
                            cutlist2.append(cut)
                    # 가공 위치순으로 정렬
                    list = []
                    cutting = []
                    for f in cutlist2:
                        list.append(f['CUT'])
                    s = sorted(list, key=lambda x: float(x[0]))
                    for l in s:
                        cutt = {'CUT': l}
                        cutting.append(cutt)
                    fielddata['cutlist'] = cutting


                    fieldlist.append(fielddata)  # 부재 하나씩 추가하여 부재별 리스트 만들어짐
                    fielddata = {}

            serializer = RpcagSerializer(data=fieldlist, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(Ok)
            else:
                return Response('serError')
        except:
            return Response(Error)


# ==============================================================================
#  robot_angle_cutting_machine --- work 그룹 조회
# ==============================================================================

class RpcagWorkDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(
                request.data['machineID'])  # codesys에 박아넣음
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            # standard = request.data['standard']
            try:
                work_list = RpcagMachine.objects.filter(machine_id=machine_id,
                                                        status=False)
            except RpcagMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

            tmp_dict = {}
            for obj in work_list:  # 시퀀스의 피스 합계 구하는 파트
                if obj.work_select == True:
                    sel = 'o'
                else:
                    sel = 'x'
                group = obj.group_data + '-' + sel
                tmp_dict[group] = tmp_dict.get(group, 0) + obj.work_quantity
            print(tmp_dict)
            mark_dict = {}
            list = []
            tmp = tmp_dict.keys()
            count = len(tmp)
            for k in tmp:
                mark_dict['count'] = str(count)
                k_list = k.split(',')
                mark_dict['group_data'] = k_list[0]
                mark_dict['work_quantity'] = str(tmp_dict[k])
                mark_dict['work_select'] = k_list[1]  # str 임
                list.append(mark_dict)
                mark_dict = {}  # 비워주지 않으면 마지막 데이터로 리스트가 다 채워짐

            return Response(list)
        except:
            return Response(Error)


# ==============================================================================
# 소요 자재 리스트 보기 ---- rpcag req material
# ==============================================================================

class RpcagReqMaterial(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = request.data['machineID']
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            # standard = request.data['standard']
            # texture = request.data['texture']

            fieldlist = []
            try:
                work_list = RpcagMachine.objects.filter(machine_id=machine_id,
                                                       status=False,
                                                       work_select=True)
            except RpcagMachine.DoesNotExist:
                return Response(status=status.HTTP_404_NOT_FOUND)

            tmp_dict = {}
            for obj in work_list:  # 시퀀스의 피스 합계 구하는 파트
                group = obj.standard + ',' + obj.texture
                tmp_dict[group] = tmp_dict.get(group, 0) + obj.length_cut
            print(tmp_dict)
            mark_dict = {}
            list = []
            tmp = tmp_dict.keys()
            count = len(tmp)
            for k in tmp:
                mark_dict['count'] = str(count)
                k_list = k.split(',')
                mark_dict['standard'] = k_list[0]
                mark_dict['texture'] = k_list[1]
                mark_dict['length'] = str(tmp_dict[k])
                list.append(mark_dict)
                mark_dict = {}  # 비워주지 않으면 마지막 데이터로 리스트가 다 채워짐

            return Response(list)
        except:
            return Response(Error)


# ==============================================================================
#  robot_angle_cutting_machine --- Part(view) 데이터 내려받기
# ==============================================================================

class RpcagPartDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = request.data['machineID']
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            standard = request.data['standard']
            texture = request.data['texture']
            inputlength = int(request.data['inputlength'])
            if inputlength < 6000:
                return Response(Error)

            # ---------------------------------------------------------------
            try:
                partdatas = RpcagMachine.objects.filter(machine_id=machine_id,
                                                        status=False,
                                                        work_select=True,
                                                        texture=texture,
                                                        standard=standard)
            except RpcagMachine.DoesNotExist:
                return Response(Error)

            # 먼저 해당 부재의 수량 만큼 배열로 정렬한다.
            list = []
            partdict = {}

            for part in partdatas:
                partcount = part.work_quantity - part.worked_quantity
                for i in range(1, partcount + 1):
                    partdict['material'] = part.view_data
                    partdict['length'] = part.length_cut
                    partdict['cutPoint'] = part.part_point
                    list.append(partdict)
                    partdict = {}

            # 배열의 인덱스를 돌면서 가공조합을 만든다.

            totallen = 0  # 가공부재 길이
            difflen = 0  # 투입부재 잔량
            endminlen = 1300  # 마지막 부재의 최소길이
            endok = False  # 마지막 부재 있을때 True
            partlist = []
            count = 0

            for index, value in enumerate(list):
                totallen = totallen + value['length']
                difflen = inputlength - totallen

                if difflen >= endminlen:  # 길이가 많이 남음
                    if value['length'] > endminlen:
                        endok = True

                elif difflen < endminlen:
                    if difflen < 0:
                        totallen = totallen - value['length']
                        break
                    else:
                        if value['length'] > endminlen:
                            endok = True

                partlist.append(value)
                del list[index]

            # 응답하는 부재의 개수를 추가한다

            lastcount = len(partlist)
            lastdict = {}
            lastlist = []
            for part in partlist:
                lastdict['count'] = str(lastcount)
                lastdict['material'] = part['material']
                lastdict['length'] = str(part['length'])
                lastdict['cutPoint'] = str(part['cutPoint'])
                lastlist.append(lastdict)
                lastdict = {}

            return Response(totallen)

        except:
            return Response(Error)

# ==============================================================================
#  robot_angle_cutting_machine --- 작업 그룹 개별 작업 선택
# ==============================================================================

class RpcagGselect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            group_data = request.data['group_data']
            texture = request.data['texture']

            try:
                select_fields = RpcagMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            texture=texture,
                                                            group_data=group_data)
                select_fields.update(work_select=True)
                return Response(Ok)

            except RpcagMachine.DoesNotExist:
                return Response(Error)

        except:
            return Response(Error)

# ==============================================================================
#  robot_angle_cutting_machine --- 전체 그룹 작업 선택
# ==============================================================================

class RpcagASelect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            # standard = request.data['standard']

            try:
                select_fields = RpcagMachine.objects.filter(machine_id=machine_id,
                                                             status=False)
                select_fields.update(work_select=True)
                return Response(Ok)

            except RpcagMachine.DoesNotExist:
                return Response(Error)

        except:
            return Response(Error)

# ==============================================================================
#  robot_angle_cutting_machine --- 가공 그룹 선택 해제
# ==============================================================================

class RpcagAcancle(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            # standard = request.data['standard']

            try:
                select_fields = RpcagMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True)
                select_fields.update(work_select=False)
                return Response(Ok)

            except RpcagMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

        except:
            return Response(Error)

# ==============================================================================
#  robot_angle_cutting_machine --- 가공 데이터 내려받기
# ==============================================================================

class RpcagCutDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            view_data = request.data['view_data']
            # standard = request.data['standard']

            try:
                cutdata = RpcagMachine.objects.get(machine_id=machine_id,
                                                    status=False,
                                                    work_select=True,
                                                    view_data=view_data)
                cutjson = cutdata.cutlist
                print(type(cutjson))
                m_dict = {}
                cutlist = []
                for cut in cutjson:
                    m_dict['count'] = cutdata.part_point
                    m_dict['pointDist'] = cut['CUT'][0]
                    m_dict['macroName'] = cut['CUT'][1]
                    m_dict['cutParam1'] = cut['CUT'][2]
                    m_dict['cutParam2'] = cut['CUT'][3]
                    m_dict['cutParam3'] = cut['CUT'][4]
                    m_dict['cutParam4'] = cut['CUT'][5]
                    cnt = len(cut['CUT'])
                    if cnt == 6:
                        m_dict['cutParam5'] = '0'
                    else:
                        m_dict['cutParam5'] = cut['CUT'][6]

                    cutlist.append(m_dict)
                    m_dict = {}
                return Response(cutlist)

            except RpcagMachine.DoesNotExist:
                return Response(Error)
        except:
            return Response(Error)

# ==============================================================================
#  robot_angle_cutting_machine --- 가공 실적 등록
# ==============================================================================

class RpcagWorkedData(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            view_data = request.data['view_data']
            # standard = request.data['standard']

            try:
                partdatas = RpcagMachine.objects.filter(machine_id=machine_id,
                                                        status=False,
                                                        work_select=True,
                                                        view_data=view_data)
                for m in partdatas:
                    if m.work_quantity > m.worked_quantity:
                        m.worked_quantity = m.worked_quantity + 1
                        if m.work_quantity == m.worked_quantity:
                            m.status = True
                    else:
                        m.status = True
                    m.save()

                return Response(Ok)

            except RpcagMachine.DoesNotExist:
                return Response(Error)
        except:
            return Response(Error)


# ==============================================================================
#  rpcag material spec  -- 자재 정보 받기
# ==============================================================================

class RpcagMaterialSpec(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            m_kinds = request.data['standard']

            # ---------단위중량, x,   y,   z,   rx,  ry,  rz,상부각,하부각,높이,중심점
            # 추후 비젼으로 대체
            origin = ['1', '1', '1', '1', '1', '1']
            # 계산용 초기값
            res = ['1', '1', '1', '1', '1', '1', '1', '1', '1', '1', '1']

            # pos = origin.split(',')  # 장비 원점 좌표

            kind = m_kinds.split(' ')  # 자재종류
            size = kind[1].split('*')   # 자재규격

            if machine_id == 2:
                res[0] = '10'
                res[1] = origin[0]  # 절대원점 x 그대로 대입
                res[2] = origin[1]  # 절대원점 y 그대로 대입
                a = float(size[0]) * round(math.radians(45), 1)  # 앵글 삼각형 중심선 높이
                b = size[2].replace('T', '')
                c = float(b[0]) * round(math.radians(45), 1)  # 두께에 대한 꼭지점 높이
                res[3] = str(round(float(origin[2]) + a + (c * 0.8), 1))
                res[4] = origin[3]  # 절대원점 rx 그대로 대입
                res[5] = origin[4]  # 절대원점 ry 그대로 대입
                res[6] = origin[5]  # 절대원점 rz 그대로 대입
                res[7] = '45'
                res[8] = '45'
                res[9] = res[3]  # 자재원점 z 대입
                res[9] = origin[1]  # 절대원점 y 그대로 대입

                return Response(res)
            else:
                return Response(Error)
        except:
            return Response(Error)


# ==============================================================================
#  rpcag speed spec  -- 자재 정보 받기
# ==============================================================================

class RpcagSpeedSpec(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            standard = request.data['standard']
            # texture = request.data['texture']
            print('1')
            with open('code_master/_robot_func/_speed_agcut.json', 'r') as f:
                json_data = json.load(f)

            return Response(json_data[standard])
        except:
            return Response(Error)


# ==============================================================================
# rpcag_camshot --- 앵글가공기 비젼
# ==============================================================================

class RpcagCamshot(APIView):
    def post(self, request, *args, **kwargs):
        try:
            origin_image = request.FILES['image']
            m_data = request.data['m_data']
            MachineKey = request.data['machineKey']
            if MachineKey != 'smart-robot-007':
                return Response('Error')
            req_data = 'OK,' + m_data
            # data_list = m_data.split(',')
            # data_list.insert(0, 'OK')
            return Response(req_data)

        except:
            return Response('Error')




# ==============================================================================
# ==============================================================================
# press_machine --- cutting view (get)
# ==============================================================================
# ==============================================================================


class PressCuttingView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            machine_id = int(request.GET['machineID'])

            view_data = request.GET['material']
            standard = request.GET['standard']

            cutlist = []
            try:
                cutdata = AutoPressMachine.objects.get(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True,
                                                            standard=standard,
                                                            view_data=view_data)
                cutlist = cutdata.cutlist
                print(cutlist)
                standard = cutdata.standard
                print(standard)
                length = cutdata.length_dwg
            except AutoPressMachine.DoesNotExist:
                return Response(Error)

            return render(request, 'code_master/AutoPressMachine/press_cutting_view.html',
                          {'cutlist': cutlist, 'standard': standard, 'length': length})

        except:
            return Response('codeError')


# ==============================================================================
# press_machine --- excel upload
# ==============================================================================

class UploadPressExcelData(APIView):
    def post(self, request, *args, **kwargs):

        try:
            file = request.FILES['file_excel']
            machine_id = request.data['extra_machineID']
            # data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
            wb = load_workbook(file, read_only=True)
            sheet = wb.worksheets[0]

            author = request.user.id
            key_list = ['temp', 'ship_no', 'por_no', 'seq_no', 'block_no', 'pcs_no',
                        'part_no', 'standard', 'length_dwg', 'work_quantity', 'cutlist']
            data_list = []
            low_count = 0
            for row in sheet.rows:
                if low_count == 0:
                    low_count = low_count + 1  # 제목행 pass
                    continue

                td = {}
                count = 0  # No 컬럼 건너뛰기
                #-------------------------------------
                for cell in row:
                    td[key_list[count]] = str(cell.value)
                    if td[key_list[count]] == 'None':
                        td[key_list[count]] = '-'
                    count = count + 1
                # -------------------------------------
                cutstr = td['cutlist']
                cutstr1 = cutstr[:-2]
                cutstr2 = cutstr1[1:]
                list1 = cutstr2.split('}')
                count = len(list1)
                list2 = []
                for cut in list1:
                    cut1 = cut.split(':')
                    list2.append(cut1[1])
                clist = []
                cutlist = []
                cutdict = {}
                for c in list2:
                    c1 = c[1:]
                    c2 = c1[:-1]
                    clist = c2.split(',')
                    cutdict['CUT'] = clist
                    cutlist.append(cutdict)
                    cutdict= {}
                td['cutlist'] = cutlist
                # 마지막 앤드컷 정보를 읽는다
                macroName = cutlist[count-1]['CUT'][1]
                cutParam = cutlist[count-1]['CUT'][2]

                # End Cut인 경우 왼쪽 수직라인 부분이 컷팅 치수의 기준라인 임
                if macroName == 'FBP02':
                    td['length_cut'] = int(td['length_dwg']) + 15
                elif macroName == 'FBP03':
                    if cutParam == '22':
                        td['length_cut'] = int(td['length_dwg']) + 15
                    elif cutParam == '25A':
                        td['length_cut'] = int(td['length_dwg']) + 10
                    elif cutParam == '32A':
                        td['length_cut'] = int(td['length_dwg']) + 5
                elif macroName == 'FBP04':
                    td['length_cut'] = int(td['length_dwg']) + 52
                else:
                    return Response("m_Error")
                # -------------------------------

                # view 데이터 자동생성
                td['view_data'] = str(td['ship_no']) + '-' + str(td['por_no']) + '-' + str(td['seq_no']) + '-' \
                                  + str(td['block_no']) + '-' + str(td['pcs_no']) + '-' + str(td['part_no'])
                td['work_quantity'] = int(td['work_quantity'])
                td['length_dwg'] = int(td['length_dwg'])
                td['part_point'] = count
                td['author'] = author
                td['machine_id'] = machine_id
                del td['temp']  # No 항목 삭제

                data_list.append(td)
                td = {}
            wb.close()
            # print(data_list)
            serializer = AutoPressSerializer(data=data_list, many=True)
            if serializer.is_valid():
                serializer.save()
                return Response(Ok)
            else:
                print(serializer.errors)
                return Response(serializer.errors,
                                status=status.HTTP_400_BAD_REQUEST)
                # return Response("s_Error")
        except:
            return Response("e_Error")


# ==============================================================================
# press_machine ---  work 그룹 조회
# ==============================================================================

class PressWorkDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID']) #codesys에 박아넣음
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            standard = request.data['standard']
            try:
                work_list = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            standard=standard)[:20]
            except AutoPressMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

            count = work_list.count()

            m_dict = {}
            list = []
            for obj in work_list:
                m_dict['count'] = str(count)
                m_dict['view_data'] = obj.view_data
                m_dict['work_quantity'] = str(obj.work_quantity)
                m_dict['worked_quantity'] = str(obj.worked_quantity)
                if obj.work_select == True:
                    select = 'o'
                else:
                    select = 'x'
                m_dict['work_select'] = select
                list.append(m_dict)
                print(m_dict)
                m_dict = {}  # 비워주지 않으면 마지막 데이터로 리스트가 다 채워짐

            return Response(list)
        except:
            return Response(Error)


# ==============================================================================
# press_machine ---  Part(view) 데이터 내려받기
# ==============================================================================

class PressPartDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = request.data['machineID']
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            standard = request.data['standard']
            inputlength = int(request.data['inputlength'])
            if inputlength < 1300:
                return Response(Error)

            # ---------------------------------------------------------------
            try:
                partdatas = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                work_select=True,
                                                                standard=standard)
            except AutoPressMachine.DoesNotExist:
                return Response(Error)

            # 먼저 해당 부재의 수량 만큼 배열로 정렬한다.
            partdict = {}
            totallen = 0  # 가공부재 길이
            difflen =0  # 투입부재 잔량
            endminlen = 700  # 마지막 부재의 최소길이
            endok = False  # 마지막 부재 있을때 True
            endng = False  # 정렬의 마지막 부재가 최소 길이 이하인지 체크
            endngcount = 0
            partlist = []

            for part in partdatas:
                partcount = part.work_quantity - part.worked_quantity
                for i in range(1, partcount + 1):
                    totallen = totallen + part.length_cut
                    difflen = inputlength - totallen
                    # 길이가 많이 남았을때 끝단 최소 길이보다 큰것이 들어가는지 체크
                    if difflen >= endminlen:
                        if part.length_cut > endminlen:
                            endok = True
                    # 길이가 끝단 최소길이보다 작게 남을때
                    elif difflen < endminlen:
                        # 부재 길이가 가공 범위를 초과 할때 길이의 합에서 마지막 더한 가공 길이 차감후 다음 스탭으로
                        if difflen < 0:
                            totallen = totallen - part.length_cut
                            continue
                        # 부재 길이가 가공 범위를 초과 하지 않을때
                        else:
                            if part.length_cut > endminlen:
                                endok = True
                                endng = False
                            # 마지막 부재가 끝단 최소 길이보다 작게 반복 루틴을 마무리 하는지 체크
                            else:
                                endng = True
                                endngcount = endngcount + 1

                    partdict['material'] = part.view_data
                    partdict['length'] = part.length_cut
                    partdict['cutPoint'] = part.part_point
                    partlist.append(partdict)
                    partdict = {}

            # 정렬의 마지막 부재가 끝단 최소 길이보다 작을시 작은수만큼 꺼내서 맨앞으로 보내줌
            if endng:
                for i in range(1, endngcount + 1):
                    end_m = partlist.pop(-1)
                    partlist.insert(0, end_m)

            lastcount = len(partlist)
            lastdict = {}
            lastlist = []
            for part in partlist:
                lastdict['count'] = str(lastcount)
                lastdict['material'] = part['material']
                lastdict['length'] = str(part['length'])
                lastdict['cutPoint'] = str(part['cutPoint'])
                lastlist.append(lastdict)
                lastdict = {}

            return Response(lastlist)

        except:
            return Response(Error)


# ==============================================================================
# press_machine_optimize ---  Part(view) 최적화 데이터 내려받기
# ==============================================================================

class PressPartDataOptimize(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = request.data['machineID']
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            standard = request.data['standard']
            inputlength = int(request.data['inputlength'])
            if inputlength < 1300:
                return Response(Error)

            # ---------------------------------------------------------------
            try:
                partdatas = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                work_select=True,
                                                                standard=standard)
            except AutoPressMachine.DoesNotExist:
                return Response(Error)

            # 먼저 해당 부재의 수량 만큼 배열로 정렬한다.
            partdict = {}
            totallen = 0  # 가공부재 길이
            difflen =0  # 투입부재 잔량
            endminlen = 1000  # 마지막 부재의 최소길이
            minimumlen = 700  # 가공 부재 최소 길이(정렬에서 제외)
            endok = False  # 마지막 부재 있을때 True
            endng = False  # 정렬의 마지막 부재가 최소 길이 이하인지 체크
            endngcount = 0
            partlist = []

            for part in partdatas:
                partcount = part.work_quantity - part.worked_quantity
                for i in range(1, partcount + 1):
                    totallen = totallen + part.length_cut
                    difflen = inputlength - totallen
                    # 길이가 최소 가공 부재 길이보다 작을시 Pass 시킴
                    if part.length_cut > minimumlen: 
                        continue
                    # 길이가 많이 남았을때 끝단 최소 길이보다 큰것이 들어가는지 체크
                    if difflen >= endminlen:
                        if part.length_cut > endminlen:
                            endok = True
                    # 길이가 끝단 최소길이보다 작게 남을때
                    elif difflen < endminlen:
                        # 부재 길이가 가공 범위를 초과 할때 길이의 합에서 마지막 더한 가공 길이 차감후 다음 스탭으로
                        if difflen < 0:
                            totallen = totallen - part.length_cut
                            continue
                        # 부재 길이가 가공 범위를 초과 하지 않을때
                        else:
                            if part.length_cut > endminlen:
                                endok = True
                                endng = False
                            # 마지막 부재가 끝단 최소 길이보다 작게 반복 루틴을 마무리 하는지 체크
                            else:
                                endng = True
                                endngcount = endngcount + 1

                    partdict['material'] = part.view_data
                    partdict['length'] = part.length_cut
                    partdict['cutPoint'] = part.part_point
                    partlist.append(partdict)
                    partdict = {}

            # 정렬의 마지막 부재가 끝단 최소 길이보다 작을시 작은수만큼 꺼내서 맨앞으로 보내줌
            if endng:
                for i in range(1, endngcount + 1):
                    end_m = partlist.pop(-1)
                    partlist.insert(0, end_m)

            lastcount = len(partlist)
            lastdict = {}
            lastlist = []
            for part in partlist:
                lastdict['count'] = str(lastcount)
                lastdict['material'] = part['material']
                lastdict['length'] = str(part['length'])
                lastdict['cutPoint'] = str(part['cutPoint'])
                lastlist.append(lastdict)
                lastdict = {}

            return Response(lastlist)

        except:
            return Response(Error)


# ==============================================================================
# press_machine ---  작업 그룹 개별 선택
# ==============================================================================

class PressGselect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            view_data = request.data['view_data']
            standard = request.data['standard']

            try:
                select_fields = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                view_data=view_data,
                                                                standard=standard)
                select_fields.update(work_select=True)
                return Response(Ok)

            except AutoPressMachine.DoesNotExist:
                return Response(Error)

        except:
            return Response(Error)

# ==============================================================================
# press_machine --- 작업 그룹 전체 작업 선택
# ==============================================================================

class PressASelect(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            standard = request.data['standard']

            try:
                select_fields = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                standard=standard)
                select_fields.update(work_select=True)
                return Response(Ok)

            except AutoPressMachine.DoesNotExist:
                return Response(Error)

        except:
            return Response(Error)


# ==============================================================================
# press_machine --- 가공 그룹 선택 해제
# ==============================================================================

class PressGcancle(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            standard = request.data['standard']

            try:
                select_fields = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                                status=False,
                                                                work_select=True,
                                                                standard=standard)
                select_fields.update(work_select=False)
                return Response(Ok)

            except AutoPressMachine.DoesNotExist:
                return Response(Error)
                # return Response(status=status.HTTP_404_NOT_FOUND)

        except:
            return Response(Error)


# ==============================================================================
# press_machine ---  가공 데이터(json) 내려받기
# ==============================================================================

class PressCutDataLoad(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)
                # return Response(status=status.HTTP_400_BAD_REQUEST)

            view_data = request.data['view_data']
            standard = request.data['standard']

            try:
                cutdata = AutoPressMachine.objects.get(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True,
                                                            standard=standard,
                                                            view_data=view_data)
                cutjson = cutdata.cutlist
                print(type(cutjson))
                m_dict = {}
                cutlist = []
                for cut in cutjson:
                    m_dict['count'] = str(cutdata.part_point)
                    m_dict['pointDist'] = cut['CUT'][0]
                    m_dict['macroName'] = cut['CUT'][1]
                    m_dict['cutParam'] = cut['CUT'][2]
                    cutlist.append(m_dict)
                    m_dict = {}
                return Response(cutlist)

            except AutoPressMachine.DoesNotExist:
                return Response(Error)
        except:
            return Response(Error)


# ==============================================================================
# press_machine ---  프레스 가공 실적 등록
# ==============================================================================

class PressWorkedData(APIView):
    def post(self, request, *args, **kwargs):
        try:
            machine_id = int(request.data['machineID'])
            machineKey = request.data['machineKey']
            if machineKey != 'smart-robot-007':
                return Response(Error)

            view_data = request.data['view_data']
            standard = request.data['standard']

            try:
                partdatas = AutoPressMachine.objects.filter(machine_id=machine_id,
                                                            status=False,
                                                            work_select=True,
                                                            standard=standard,
                                                            view_data=view_data)
                for m in partdatas:
                    if m.work_quantity > m.worked_quantity:
                        m.worked_quantity = m.worked_quantity + 1
                        if m.work_quantity == m.worked_quantity:
                            m.status = True
                    else:
                        m.status = True
                    m.save()

                return Response(Ok)

            except AutoPressMachine.DoesNotExist:
                return Response(Error)
        except:
            return Response(Error)


# ==============================================================================
# ==============================================================================
# 4. Smart Cam Machine
# ==============================================================================
# ==============================================================================
from .models import SmartCamMachine
from .serializers import SmartCamSerializer

from .camshot.rpcm_s300._lib import __find as F
import cv2
import numpy as np
import requests
from django.core.files.base import ContentFile
import uuid
# ==============================================================================
# RPCM Camshot ---
# ==============================================================================

class RpcmCamshot(APIView):
    def post(self, request, *args, **kwargs):
        try:
            start = time.time()
            origin_image = request.FILES['image']
            # cam_name = request.data['cam_name']
            m_data = request.data['m_data']
            MachineKey = request.data['machineKey']
            if MachineKey != 'smart-robot-007':
                return Response('NG,key_Error')

            data_list = m_data.split(',')
            cam_name = f'miju_rpcm_s300_{data_list[1]}'
            kind = data_list[1].split(' ')

            # 업로드된 데이터 먼저 저장후 영상 처리 할때 적용 할 것(삭제금지)-----------------------------
            uploaddata = {'cam_name': cam_name, 'origin_image': origin_image}
            serializer = SmartCamSerializer(data=uploaddata)
            if serializer.is_valid():
                serializer.save()
            else:
                return Response('NG,Serializer_Error')

            query = SmartCamMachine.objects.filter(cam_name= cam_name).first()
            img_path = query.origin_image
            # url을 통해서 이미지 읽어 올때 ----------------------------------------
            image_nparray = np.asarray(bytearray(requests.get(img_path.url).content), dtype=np.uint8)
            image = cv2.imdecode(image_nparray, cv2.IMREAD_GRAYSCALE)
            # 인터넷에 다른 방법이 있길래 참고 삼아 적어둠
            # image = cv2.imdecode(numpy.frombuffer(myfile, numpy.uint8), cv2.IMREAD_UNCHANGED)
            # ------------------------------------------------------------------

            # 비젼 처리 파트 ---
            # =========================================================================================
            # >>> 호출하는 부분 ========================================================================
            # =========================================================================================

            _TYPE = kind[0]
            obj = F.MyClass(_TYPE)
            (result, dst,(topX, topY, middleX, middleY, bottomX, bottomY)) = obj.Start(data_list[1], image)

            # <<< =====================================================================================

            # sim_point 초기화 --------------------------------------------------
            sim_point = {}
            sim_point["cmd"] = 'OK'
            sim_point["WorkSize"] = data_list[1]
            sim_point["StanValueX"] = data_list[2]
            sim_point["StanValueY"] = data_list[3]
            sim_point["StanValueZ"] = data_list[4]
            sim_point["StanAdjustX"] = data_list[5]
            sim_point["StanAdjustY"] = data_list[6]
            sim_point["StanAdjustZ"] = data_list[7]
            sim_point["SizeAdjustA"] = data_list[8]
            sim_point["SizeAdjustB"] = data_list[9]
            sim_point["SizeAdjustC"] = data_list[10]
            sim_point["SizeAdjustD"] = data_list[11]
            sim_point["BeamAdjustMH"] = data_list[12]
            sim_point['result_data'] = [result, topX, topY, middleX, middleY,
                                        bottomX, bottomY]

            # cam_data(해당 자재의 영상인식 평균 데이터) 만들기
            cam_data = ''
            if SmartCamMachine.objects.filter(cam_name=f'{cam_name}_OK').exists():
                _query = SmartCamMachine.objects.filter(cam_name=f'{cam_name}_OK').first()
                cam_data = _query.cam_data
            else:
                cam_data = f'{topX},{topY},{middleX},{middleY},{bottomX},{bottomY}'

            # 영상 인식의 평균 데이터 (최초 실행시는 최초 영상 인식 데이터) - 현장 캘리브레이션 적용
            avr = cam_data.split(',')
            (tX, tY, mX, mY, bX, bY) = (0.0, 0.0, 0.0, 0.0, 0.0, 0.0)
            if result is None:  # 영상처리가 정상 처리 되었을때
                query.cam_name = cam_name + '_OK'
                tX = float(avr[0]) - float(topX)
                tY = float(avr[1]) - float(topY)
                mX = float(avr[2]) - float(middleX)
                mY = float(avr[3]) - float(middleY)
                bX = float(avr[4]) - float(bottomX)
                bY = float(avr[5]) - float(bottomY)
                if abs(tX) < 5 and abs(tY) < 10 and abs(mX) < 5 and abs(mY) < 10 and \
                        abs(bX) < 5 and abs(bY) < 10:
                    _tX = (float(topX) + (float(avr[0]) * 10)) / 11
                    _tY = (float(topY) + (float(avr[1]) * 10)) / 11
                    _mX = (float(middleX) + (float(avr[2]) * 10)) / 11
                    _mY = (float(middleY) + (float(avr[3]) * 10)) / 11
                    _bX = (float(bottomX) + (float(avr[4]) * 10)) / 11
                    _bY = (float(bottomY) + (float(avr[5]) * 10)) / 11
                    # cam_data = 자재 규격별 영상 처리 결과의 축척된 평균 데이터 update
                    cam_data = f'{str(_tX)},{str(_tY)},{str(_mX)},{str(_mY)},{str(_bX)},{str(_bY)}'

                    myUUID = uuid.uuid1()
                    ret, buf = cv2.imencode('.jpg', dst)  # dst: nd array to img
                    content = ContentFile(buf.tobytes())
                    query.result_image.save(f'{myUUID}.jpg', content, save=False)
                    query.sim_point= sim_point
                    query.cam_data= cam_data
                    query.save()

                else:   # 좌표 편차의 평균이 일정 값 이상일때
                    query.cam_name= cam_name + '_DIFF'

                    myUUID = uuid.uuid1()
                    ret, buf = cv2.imencode('.jpg', dst)  # dst: nd array to img
                    content = ContentFile(buf.tobytes())
                    query.result_image.save(f'{myUUID}.jpg', content, save=False)
                    query.sim_point= sim_point
                    query.cam_data= cam_data
                    query.save()

                    result_str= f'NG, 인식 편차가 심합니다. 자재 원점을 수동으로 조정하세요'
                    return Response(result_str)

            # 영상 인식 애러 발생시(추후 애러 코드 분류 및 반영 할 것)
            else:
                query.cam_name= cam_name + '_NG'

                myUUID = uuid.uuid1()
                ret, buf = cv2.imencode('.jpg', dst)  # dst: nd array to img
                content = ContentFile(buf.tobytes())
                query.result_image.save(f'{myUUID}.jpg', content, save=False)
                query.sim_point= sim_point
                cam_data = f'{topX},{topY},{middleX},{middleY},{bottomX},{bottomY}'
                query.cam_data= cam_data
                query.save()

                result_str= f'NG, 영상인식 애러 >> 자재 원점을 수동으로 조정하세요'
                return Response(result_str)
            # ---------------------------------------------------------------------<<<

            # 위 부분 삭제시 tX~bY를 topX~bottomY로 바꿀것-----------------------------|||

            # 영상인식 OK일때 보정값 계산하기------------------------------------------->>>

            # adjust 보정 좌표값 계산
            # EA 50*50*4T, UA 125*75*7/7T, PI 103A-114.3, SP 75*75
            # CH 100*50*6/8.5T, IB 150*125*8.5/14T, HB 125*125*6.5/9T
            # 스케일계산: 720(거리)- f_h * 2.4(sen_h) / 4(focal length) = 432mm(1080px)

            if result is None:
                if kind[0] == 'EA' or kind[0] == 'UA':
                    m_h = data_list[1].split('*')
                    f_h = 720 - float(m_h[1]) * 0.77
                    focal = round((f_h * 2.4 / 4 / 1080), 1)   # mm/picel
                    _aX = (tX + mX + bX)/3
                    sim_point['StanAdjustX'] = str(round((_aX * focal), 1))
                    _aY = (tY + mY + bY)/3
                    sim_point['StanAdjustY'] = str(round((_aY * focal), 1))
                elif kind[0] == 'CH' or kind[0] == 'HB' or kind[0] == 'IB':
                    m_h = data_list[1].split('*')
                    f_h = 720 - float(m_h[1])
                    focal = round((f_h * 2.4 / 4 / 1080), 1)   # mm/picel
                    sim_point['StanAdjustX'] = str(round((bX * focal), 1))
                    sim_point['StanAdjustY'] = str(round((bY * focal), 1))
                    sim_point["SizeAdjustA"] = str(round((tY * focal), 1))
                elif kind[0] == 'PI':
                    m_h = data_list[1].split('-')
                    f_h = 720 - float(m_h[1])
                    focal = round((f_h * 2.4 / 4 / 1080), 1)   # mm/picel
                    sim_point['StanAdjustX'] = str(round((mX * focal), 1))
                else:
                    pass

                stX = sim_point["StanValueX"]
                stY = sim_point["StanValueY"]
                stZ = sim_point["StanValueZ"]
                adX = sim_point["StanAdjustX"]
                adY = sim_point["StanAdjustY"]
                adZ = sim_point["StanAdjustZ"]
                szA = sim_point["SizeAdjustA"]
                szB = sim_point["SizeAdjustB"]
                szC = sim_point["SizeAdjustC"]
                szD = sim_point["SizeAdjustD"]
                bmH = sim_point["BeamAdjustMH"]
                result_str = f'{sim_point["cmd"]},{sim_point["WorkSize"]},{stX},{stY},{stZ},' \
                             f'{adX},{adY},{adZ},{szA},{szB},{szC},{szD},{bmH}'
            else:
                result_str= f'NG, 애러코드: {result} >> 자재 원점을 수동으로 조정하세요'

            # print("time :", time.time() - start)
            return Response(result_str)

        except Exception as ee:
            print(f"Unexpected {ee=}, {type(ee)=}")
            return Response('NG,except_Error')

